"""Tests for scripts/build_descriptions.py."""

import json
from pathlib import Path

import openpyxl
import pytest

from scripts.build_descriptions import (
    _clean_cell,
    _normalize_header_label,
    _text_to_bullets,
    find_header_row,
    load_definition_workbook,
    load_occupations,
    main,
    merge_definitions,
    render_description,
    write_descriptions,
)

# ── Fixtures ─────────────────────────────────────────────────────────────


def _write_occupations_json(tmp_path: Path, records: list[dict] | None = None) -> Path:
    """Write a synthetic sg_occupations.json."""
    if records is None:
        records = [
            {
                "title": "Software Developer",
                "slug": "software-developer",
                "ssoc_code": "25121",
                "category": "professionals",
                "category_label": "Professionals",
                "major_group": 2,
                "pay_monthly": 6600,
                "pay_annual": 79200,
                "pay_p25": 4900,
                "pay_p75": 9300,
                "url": "",
            },
            {
                "title": "Accountant",
                "slug": "accountant",
                "ssoc_code": "21112",
                "category": "professionals",
                "category_label": "Professionals",
                "major_group": 2,
                "pay_monthly": 5000,
                "pay_annual": 60000,
                "pay_p25": 3758,
                "pay_p75": 6700,
                "url": "",
            },
        ]
    path = tmp_path / "occupations.json"
    path.write_text(json.dumps(records), encoding="utf-8")
    return path


def _build_definition_workbook(
    tmp_path: Path,
    *,
    filename: str = "ssoc_defs.xlsx",
    sheet_name: str = "Definitions",
    header_row: int = 5,
    rows: list[tuple] | None = None,
    headers: tuple | None = None,
) -> Path:
    """Create a synthetic SSOC definitions workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if headers is None:
        headers = (
            "SSOC 2020",
            "SSOC 2020 Title",
            "Groups Classified Under this Code",
            "Detailed Definitions",
            "Tasks",
            "Notes",
            "Examples of Job Classified Under this Code",
            "Examples of Job Classified Elsewhere",
        )

    for j, h in enumerate(headers):
        ws.cell(row=header_row, column=j + 1, value=h)

    if rows is None:
        rows = [
            (
                "25121",
                "Software developer",
                None,
                "Develops software applications and systems.",
                "- designing software; coding; testing; debugging",
                None,
                "\u2022 Application developer\n\u2022 Backend developer",
                None,
            ),
            (
                "21112",
                "Accountant",
                None,
                "Prepares and examines financial records.",
                "- preparing accounts; auditing; tax advisory",
                None,
                "\u2022 Tax accountant\n\u2022 Financial accountant",
                None,
            ),
            # Non-5-digit code (should be skipped)
            (
                "2",
                "PROFESSIONALS",
                None,
                "<Blank>",
                "<Blank>",
                None,
                "<Blank>",
                None,
            ),
        ]

    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=header_row + 1 + i, column=j + 1, value=val)

    path = tmp_path / filename
    wb.save(path)
    wb.close()
    return path


def _make_occupation(*, slug: str = "test", ssoc_code: str = "12345", **overrides: object) -> dict:
    """Build a minimal valid occupation record."""
    base = {
        "title": "Test Occupation",
        "slug": slug,
        "ssoc_code": ssoc_code,
        "category": "professionals",
        "category_label": "Professionals",
        "major_group": 2,
        "pay_monthly": 5000,
        "pay_annual": 60000,
        "pay_p25": 4000,
        "pay_p75": 6000,
        "url": "",
    }
    base.update(overrides)
    return base


# ── Tests: _normalize_header_label ────────────────────────────────


class TestNormalizeHeaderLabel:
    def test_lowercase(self) -> None:
        assert _normalize_header_label("SSOC 2020") == "ssoc 2020"

    def test_ampersand(self) -> None:
        assert _normalize_header_label("Trade & Industry") == "trade and industry"

    def test_whitespace(self) -> None:
        assert _normalize_header_label("  Detailed   Definitions  ") == "detailed definitions"


# ── Tests: _clean_cell ────────────────────────────────────────────


class TestCleanCell:
    def test_none(self) -> None:
        assert _clean_cell(None) is None

    def test_blank_marker(self) -> None:
        assert _clean_cell("<Blank>") is None

    def test_empty(self) -> None:
        assert _clean_cell("") is None

    def test_real_text(self) -> None:
        assert _clean_cell("Develops software.") == "Develops software."

    def test_special_chars(self) -> None:
        result = _clean_cell("\u2022\x9f\xa0 Item one")
        assert result is not None
        assert "Item one" in result


# ── Tests: _text_to_bullets ───────────────────────────────────────


class TestTextToBullets:
    def test_semicolons(self) -> None:
        result = _text_to_bullets("designing; coding; testing")
        assert result.startswith("- ")
        assert result.count("\n") == 2

    def test_single_item_stays_paragraph(self) -> None:
        text = "A single long sentence describing tasks."
        assert _text_to_bullets(text) == text

    def test_newline_split(self) -> None:
        result = _text_to_bullets("task one\ntask two")
        assert "- task one" in result
        assert "- task two" in result


# ── Tests: find_header_row ────────────────────────────────────────


class TestFindHeaderRow:
    def test_standard_layout(self, tmp_path: Path) -> None:
        wb_path = _build_definition_workbook(tmp_path)
        wb = openpyxl.load_workbook(wb_path)
        ws = wb.active
        row_num, col_map = find_header_row(ws)
        assert row_num == 5
        assert "ssoc_code" in col_map
        assert "definition" in col_map
        wb.close()

    def test_missing_columns(self, tmp_path: Path) -> None:
        wb_path = _build_definition_workbook(tmp_path, headers=("Col A", "Col B", "Col C"))
        wb = openpyxl.load_workbook(wb_path)
        ws = wb.active
        with pytest.raises(ValueError, match="required columns"):
            find_header_row(ws)
        wb.close()

    def test_alternate_aliases(self, tmp_path: Path) -> None:
        wb_path = _build_definition_workbook(
            tmp_path,
            headers=("Code", "Title", None, "Occupation Definition"),
        )
        wb = openpyxl.load_workbook(wb_path)
        ws = wb.active
        _, col_map = find_header_row(ws)
        assert col_map["ssoc_code"] == 0
        assert col_map["definition"] == 3
        wb.close()


# ── Tests: load_definition_workbook ───────────────────────────────


class TestLoadDefinitionWorkbook:
    def test_basic(self, tmp_path: Path) -> None:
        wb_path = _build_definition_workbook(tmp_path)
        defs = load_definition_workbook(wb_path, source_year=2020)
        assert "25121" in defs
        assert "21112" in defs
        # Non-5-digit code should be excluded
        assert "00002" not in defs

    def test_definition_content(self, tmp_path: Path) -> None:
        wb_path = _build_definition_workbook(tmp_path)
        defs = load_definition_workbook(wb_path, source_year=2020)
        assert defs["25121"]["definition"] == "Develops software applications and systems."
        assert defs["25121"]["source_year"] == 2020

    def test_missing_workbook(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            load_definition_workbook(tmp_path / "missing.xlsx", source_year=2020)

    def test_no_valid_sheet(self, tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        ws.cell(row=1, column=1, value="nothing")
        path = tmp_path / "bad.xlsx"
        wb.save(path)
        wb.close()
        with pytest.raises(ValueError, match="required columns"):
            load_definition_workbook(path, source_year=2020)

    def test_duplicate_code_prefers_more_complete(self, tmp_path: Path) -> None:
        rows = [
            ("25121", "Dev", None, "Short def.", None, None, None, None),
            ("25121", "Dev", None, "Better def.", "Has tasks", None, "Has examples", None),
        ]
        wb_path = _build_definition_workbook(tmp_path, rows=rows)
        defs = load_definition_workbook(wb_path, source_year=2020)
        assert defs["25121"]["definition"] == "Better def."


# ── Tests: merge_definitions ──────────────────────────────────────


class TestMergeDefinitions:
    def test_both_none(self) -> None:
        assert merge_definitions(None, None) is None

    def test_primary_only(self) -> None:
        p = {"definition": "Primary def", "tasks": "Tasks", "examples": None}
        result = merge_definitions(p, None)
        assert result["definition"] == "Primary def"

    def test_supplemental_only(self) -> None:
        s = {"definition": "Sup def", "tasks": None, "examples": "Exs"}
        result = merge_definitions(None, s)
        assert result["definition"] == "Sup def"

    def test_fill_only_no_override(self) -> None:
        p = {"definition": "Primary def", "tasks": None, "examples": None}
        s = {"definition": "Sup def", "tasks": "Sup tasks", "examples": "Sup exs"}
        result = merge_definitions(p, s)
        assert result is not None
        assert result["definition"] == "Primary def"  # not overridden
        assert result["tasks"] == "Sup tasks"  # filled
        assert result["examples"] == "Sup exs"  # filled


# ── Tests: render_description ─────────────────────────────────────


class TestRenderDescription:
    def test_full_record(self) -> None:
        occ = _make_occupation(title="Software Developer", ssoc_code="25121")
        defn = {
            "definition": "Develops software.",
            "tasks": "Coding; testing",
            "examples": "App dev",
        }
        md = render_description(occ, defn)
        assert "# Software Developer (SSOC 25121)" in md
        assert "Develops software." in md
        assert "S$5,000" in md
        assert "Professionals" in md

    def test_no_definition_fallback(self) -> None:
        occ = _make_occupation()
        md = render_description(occ, None)
        assert "No official SSOC definition" in md
        assert "Not available" in md

    def test_pay_range_both(self) -> None:
        occ = _make_occupation(pay_p25=4000, pay_p75=6000)
        md = render_description(occ, None)
        assert "S$4,000 \u2013 S$6,000" in md

    def test_pay_range_p25_only(self) -> None:
        occ = _make_occupation(pay_p25=4000, pay_p75=None)
        md = render_description(occ, None)
        assert "25th percentile" in md

    def test_pay_range_p75_only(self) -> None:
        occ = _make_occupation(pay_p25=None, pay_p75=6000)
        md = render_description(occ, None)
        assert "75th percentile" in md

    def test_pay_range_neither(self) -> None:
        occ = _make_occupation(pay_p25=None, pay_p75=None)
        md = render_description(occ, None)
        assert "not available" in md


# ── Tests: load_occupations ───────────────────────────────────────


class TestLoadOccupations:
    def test_basic(self, tmp_path: Path) -> None:
        path = _write_occupations_json(tmp_path)
        records = load_occupations(path)
        assert len(records) == 2

    def test_missing_file(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            load_occupations(tmp_path / "missing.json")

    def test_not_array(self, tmp_path: Path) -> None:
        path = tmp_path / "bad.json"
        path.write_text('{"not": "array"}')
        with pytest.raises(ValueError, match="top-level array"):
            load_occupations(path)

    def test_missing_field(self, tmp_path: Path) -> None:
        records = [_make_occupation()]
        del records[0]["slug"]
        path = _write_occupations_json(tmp_path, records)
        with pytest.raises(ValueError, match="slug"):
            load_occupations(path)

    def test_duplicate_slug(self, tmp_path: Path) -> None:
        records = [
            _make_occupation(slug="same", ssoc_code="11111"),
            _make_occupation(slug="same", ssoc_code="22222"),
        ]
        path = _write_occupations_json(tmp_path, records)
        with pytest.raises(ValueError, match="duplicate slug"):
            load_occupations(path)


# ── Tests: write_descriptions ─────────────────────────────────────


class TestWriteDescriptions:
    def test_end_to_end(self, tmp_path: Path) -> None:
        occs = [
            _make_occupation(slug="dev", ssoc_code="25121", title="Developer"),
            _make_occupation(slug="acct", ssoc_code="21112", title="Accountant"),
        ]
        defs_2020 = {
            "25121": {
                "definition": "Develops software.",
                "tasks": "Coding",
                "examples": "App dev",
            }
        }
        defs_2024: dict[str, dict] = {}
        out = tmp_path / "pages"

        stats = write_descriptions(occs, defs_2020, defs_2024, out)
        assert stats["written"] == 2
        assert stats["with_2020"] == 1
        assert stats["fallback_only"] == 1
        assert (out / "dev.md").exists()
        assert (out / "acct.md").exists()

        dev_content = (out / "dev.md").read_text()
        assert "Developer" in dev_content
        assert "Develops software." in dev_content

    def test_2024_fill_tracked(self, tmp_path: Path) -> None:
        occs = [_make_occupation(slug="x", ssoc_code="11111")]
        defs_2020 = {"11111": {"definition": "Def.", "tasks": None, "examples": None}}
        defs_2024 = {"11111": {"definition": None, "tasks": "Tasks from 2024", "examples": None}}
        out = tmp_path / "pages"

        stats = write_descriptions(occs, defs_2020, defs_2024, out)
        assert stats["with_2024_fill"] == 1


# ── Tests: main (CLI) ─────────────────────────────────────────────


class TestMain:
    def test_full_run(self, tmp_path: Path) -> None:
        occ_path = _write_occupations_json(tmp_path)
        wb2020 = _build_definition_workbook(tmp_path, filename="ssoc2020.xlsx")
        wb2024 = _build_definition_workbook(tmp_path, filename="ssoc2024.xlsx")
        out_dir = tmp_path / "out_pages"

        main(
            [
                "--occupations",
                str(occ_path),
                "--ssoc2020",
                str(wb2020),
                "--ssoc2024",
                str(wb2024),
                "--output-dir",
                str(out_dir),
            ]
        )

        assert (out_dir / "software-developer.md").exists()
        assert (out_dir / "accountant.md").exists()
        content = (out_dir / "software-developer.md").read_text()
        assert "SSOC 25121" in content
        assert "S$6,600" in content
