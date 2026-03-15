"""Tests for scripts/parse_wages.py."""

import json
from pathlib import Path

import openpyxl
import pytest

from scripts.parse_wages import (
    clean_title,
    dedupe_slugs,
    detect_major_group,
    format_ssoc_code,
    make_slug,
    normalize_header,
    parse_wage,
    parse_workbook,
    validate_records,
    write_csv,
    write_json,
)

# ── Helper: build a synthetic workbook ─────────────────────────────────


def _build_workbook(
    tmp_path: Path,
    *,
    occupation_rows: list[tuple] | None = None,
    sheet_name: str = "T4",
    include_headers: bool = True,
    extra_groups: list[tuple] | None = None,
) -> Path:
    """Create a minimal wages workbook matching MOM T4 layout.

    Default: 2 major groups with 3 occupations each (6 total).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Rows 1–8: headers (only row 5–6 matter but we fill placeholders)
    if include_headers:
        for r in range(1, 9):
            ws.cell(row=r, column=1, value=None)

    # Default occupation data
    if occupation_rows is None:
        rows = [
            # Row 9: major group header
            (1, 1, "MANAGERS", None, None, None, None, None, None),
            # Row 10-12: occupations
            (2, 12112, "Administration manager", 4480, 6320, 8351, 4569, 6388, 8431),
            (3, 12221, "Marketing manager", 5900, 8429, 13218, 5943, 8514, 13333),
            (4, 13430, "Aged care services manager", 6000, 7719, 9840, 6091, 7759, 9997),
            # Row 13: major group header
            (5, 2, "PROFESSIONALS", None, None, None, None, None, None),
            # Row 14-16: occupations
            (6, 21112, "Accountant", 3700, 4929, 6600, 3758, 5000, 6700),
            (7, 21123, "Auditor", 3500, 4600, 6200, 3550, 4650, 6250),
            (8, 25121, "Software developer", 4800, 6500, 9200, 4900, 6600, 9300),
        ]
    else:
        rows = occupation_rows

    if extra_groups:
        rows = list(rows) + list(extra_groups)

    start_row = 9
    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=start_row + i, column=j + 1, value=val)

    path = tmp_path / "test_wages.xlsx"
    wb.save(path)
    wb.close()
    return path


def _make_valid_records(n: int = 562, groups: int = 9) -> list[dict]:
    """Generate n valid records spread across `groups` major groups."""
    records = []
    per_group = n // groups
    remainder = n % groups
    idx = 0
    for g in range(1, groups + 1):
        count = per_group + (1 if g <= remainder else 0)
        from scripts.parse_wages import MAJOR_GROUP_MAP

        label, cat = MAJOR_GROUP_MAP[g]
        for _ in range(count):
            code = f"{g}{idx:04d}"
            records.append(
                {
                    "title": f"Occupation {idx}",
                    "slug": f"occupation-{idx}",
                    "ssoc_code": code,
                    "category": cat,
                    "category_label": label,
                    "major_group": g,
                    "pay_monthly": 3000 + idx * 10,
                    "pay_annual": (3000 + idx * 10) * 12,
                    "pay_p25": 2500 + idx * 10,
                    "pay_p75": 3500 + idx * 10,
                    "url": "",
                }
            )
            idx += 1
    return records


# ── Tests: normalize_header ────────────────────────────────────────


class TestNormalizeHeader:
    def test_strips_leading_digit(self) -> None:
        assert normalize_header("1 MANAGERS") == "MANAGERS"

    def test_strips_leading_digits_with_separator(self) -> None:
        assert normalize_header("3. ASSOCIATE PROFESSIONALS") == "ASSOCIATE PROFESSIONALS"

    def test_replaces_ampersand(self) -> None:
        assert normalize_header("SERVICE & SALES") == "SERVICE AND SALES"

    def test_collapses_whitespace(self) -> None:
        assert normalize_header("PLANT   AND   MACHINE") == "PLANT AND MACHINE"

    def test_uppercases(self) -> None:
        assert normalize_header("managers") == "MANAGERS"


# ── Tests: detect_major_group ──────────────────────────────────────


class TestDetectMajorGroup:
    def test_known_groups(self) -> None:
        assert detect_major_group("MANAGERS") == 1
        assert detect_major_group("PROFESSIONALS") == 2
        assert detect_major_group("ASSOCIATE PROFESSIONALS AND TECHNICIANS") == 3
        assert detect_major_group("CLERICAL SUPPORT WORKERS") == 4
        assert detect_major_group("SERVICE AND SALES WORKERS") == 5
        assert detect_major_group("AGRICULTURAL AND FISHERY WORKERS") == 6
        assert detect_major_group("CRAFTSMEN AND RELATED TRADES WORKERS") == 7
        assert detect_major_group("PLANT AND MACHINE OPERATORS AND ASSEMBLERS") == 8
        assert detect_major_group("CLEANERS, LABOURERS AND RELATED WORKERS") == 9

    def test_with_leading_digit(self) -> None:
        assert detect_major_group("1 MANAGERS") == 1

    def test_unknown_returns_none(self) -> None:
        assert detect_major_group("UNKNOWN GROUP") is None
        assert detect_major_group("Software developer") is None


# ── Tests: parse_wage ─────────────────────────────────────────────


class TestParseWage:
    def test_int(self) -> None:
        assert parse_wage(4480) == 4480

    def test_float(self) -> None:
        assert parse_wage(4480.0) == 4480

    def test_string_number(self) -> None:
        assert parse_wage("4480") == 4480

    def test_string_with_commas(self) -> None:
        assert parse_wage("13,218") == 13218

    def test_none(self) -> None:
        assert parse_wage(None) is None

    def test_dash(self) -> None:
        assert parse_wage("-") is None

    def test_empty_string(self) -> None:
        assert parse_wage("") is None

    def test_whitespace_string(self) -> None:
        assert parse_wage(" ") is None

    def test_garbage(self) -> None:
        assert parse_wage("n/a") is None


# ── Tests: make_slug ──────────────────────────────────────────────


class TestMakeSlug:
    def test_basic(self) -> None:
        assert make_slug("Software developer") == "software-developer"

    def test_ampersand(self) -> None:
        assert make_slug("Food & Beverage Servers") == "food-and-beverage-servers"

    def test_parentheses(self) -> None:
        assert make_slug("Cleaners (Commercial)") == "cleaners-commercial"

    def test_apostrophe(self) -> None:
        assert make_slug("Ship's officer") == "ships-officer"

    def test_slashes(self) -> None:
        assert make_slug("Advertising/Public relations manager") == (
            "advertising-public-relations-manager"
        )

    def test_multiple_special_chars(self) -> None:
        assert make_slug("A -- B") == "a-b"


# ── Tests: dedupe_slugs ───────────────────────────────────────────


class TestDedupeSlugs:
    def test_no_dupes(self) -> None:
        records = [{"slug": "a"}, {"slug": "b"}, {"slug": "c"}]
        dedupe_slugs(records)
        assert [r["slug"] for r in records] == ["a", "b", "c"]

    def test_with_dupes(self) -> None:
        records = [{"slug": "x"}, {"slug": "x"}, {"slug": "x"}]
        dedupe_slugs(records)
        assert [r["slug"] for r in records] == ["x", "x-2", "x-3"]

    def test_mixed(self) -> None:
        records = [{"slug": "a"}, {"slug": "b"}, {"slug": "a"}]
        dedupe_slugs(records)
        assert [r["slug"] for r in records] == ["a", "b", "a-2"]


# ── Tests: clean_title ────────────────────────────────────────────


class TestCleanTitle:
    def test_collapses_whitespace(self) -> None:
        assert clean_title("A   B") == "A B"

    def test_strips_edges(self) -> None:
        assert clean_title("  hello  ") == "hello"

    def test_newlines(self) -> None:
        assert clean_title("line1\nline2") == "line1 line2"


# ── Tests: format_ssoc_code ───────────────────────────────────────


class TestFormatSsocCode:
    def test_from_int(self) -> None:
        assert format_ssoc_code(12112) == "12112"

    def test_from_small_int(self) -> None:
        assert format_ssoc_code(1234) == "01234"

    def test_from_string(self) -> None:
        assert format_ssoc_code("25121") == "25121"

    def test_from_float(self) -> None:
        assert format_ssoc_code(25121.0) == "25121"


# ── Tests: parse_workbook (with synthetic workbook) ────────────────


class TestParseWorkbook:
    def test_basic_parse(self, tmp_path: Path) -> None:
        wb_path = _build_workbook(tmp_path)
        records = parse_workbook(wb_path)
        assert len(records) == 6
        assert records[0]["title"] == "Administration manager"
        assert records[0]["ssoc_code"] == "12112"
        assert records[0]["category"] == "managers"
        assert records[0]["major_group"] == 1

    def test_gross_wages_used(self, tmp_path: Path) -> None:
        wb_path = _build_workbook(tmp_path)
        records = parse_workbook(wb_path)
        # First occupation: gross p25=4569, gross med=6388, gross p75=8431
        assert records[0]["pay_monthly"] == 6388
        assert records[0]["pay_p25"] == 4569
        assert records[0]["pay_p75"] == 8431
        assert records[0]["pay_annual"] == 6388 * 12

    def test_two_groups_found(self, tmp_path: Path) -> None:
        wb_path = _build_workbook(tmp_path)
        records = parse_workbook(wb_path)
        groups = {r["major_group"] for r in records}
        assert groups == {1, 2}

    def test_slug_generation(self, tmp_path: Path) -> None:
        wb_path = _build_workbook(tmp_path)
        records = parse_workbook(wb_path)
        assert records[0]["slug"] == "administration-manager"
        assert records[5]["slug"] == "software-developer"

    def test_missing_workbook(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            parse_workbook(tmp_path / "nonexistent.xlsx")

    def test_wrong_sheet_name(self, tmp_path: Path) -> None:
        wb_path = _build_workbook(tmp_path, sheet_name="Other")
        with pytest.raises(ValueError, match="Sheet 'T4' not found"):
            parse_workbook(wb_path, sheet_name="T4")

    def test_occupation_before_group_header(self, tmp_path: Path) -> None:
        rows = [
            # Occupation row without a preceding group header
            (1, 12112, "Admin manager", 4480, 6320, 8351, 4569, 6388, 8431),
        ]
        wb_path = _build_workbook(tmp_path, occupation_rows=rows)
        with pytest.raises(ValueError, match="before any major-group header"):
            parse_workbook(wb_path)

    def test_duplicate_title_slugs(self, tmp_path: Path) -> None:
        rows = [
            (1, 1, "MANAGERS", None, None, None, None, None, None),
            (2, 12112, "Manager", 4000, 5000, 6000, 4100, 5100, 6100),
            (3, 12113, "Manager", 4000, 5000, 6000, 4100, 5100, 6100),
        ]
        wb_path = _build_workbook(tmp_path, occupation_rows=rows)
        records = parse_workbook(wb_path)
        slugs = [r["slug"] for r in records]
        assert slugs == ["manager", "manager-2"]


# ── Tests: validate_records ───────────────────────────────────────


class TestValidateRecords:
    def test_valid_records_pass(self) -> None:
        records = _make_valid_records()
        validate_records(records)  # Should not raise

    def test_wrong_count(self) -> None:
        records = _make_valid_records(n=100)
        with pytest.raises(ValueError, match="Expected 562"):
            validate_records(records)

    def test_wrong_group_count(self) -> None:
        records = _make_valid_records()
        # Force all records to group 1
        for r in records:
            r["major_group"] = 1
        with pytest.raises(ValueError, match="Expected 9 major groups"):
            validate_records(records)

    def test_missing_title(self) -> None:
        records = _make_valid_records()
        records[0]["title"] = ""
        with pytest.raises(ValueError, match="missing title"):
            validate_records(records)

    def test_wage_ordering_violation(self) -> None:
        records = _make_valid_records()
        records[0]["pay_p25"] = 9999
        records[0]["pay_monthly"] = 1000
        with pytest.raises(ValueError, match="pay_p25"):
            validate_records(records)

    def test_duplicate_slugs(self) -> None:
        records = _make_valid_records()
        records[1]["slug"] = records[0]["slug"]
        with pytest.raises(ValueError, match="Duplicate slugs"):
            validate_records(records)

    def test_duplicate_ssoc_codes(self) -> None:
        records = _make_valid_records()
        records[1]["ssoc_code"] = records[0]["ssoc_code"]
        with pytest.raises(ValueError, match="Duplicate SSOC"):
            validate_records(records)

    def test_bad_ssoc_format(self) -> None:
        records = _make_valid_records()
        records[0]["ssoc_code"] = "123"  # Not 5 digits
        with pytest.raises(ValueError, match="not 5 digits"):
            validate_records(records)

    def test_annual_mismatch(self) -> None:
        records = _make_valid_records()
        records[0]["pay_annual"] = 999
        with pytest.raises(ValueError, match="pay_annual != pay_monthly"):
            validate_records(records)

    def test_median_gt_p75(self) -> None:
        records = _make_valid_records()
        records[0]["pay_monthly"] = 9999
        records[0]["pay_annual"] = 9999 * 12
        with pytest.raises(ValueError, match="median.*pay_p75"):
            validate_records(records)


# ── Tests: write_json / write_csv ─────────────────────────────────


class TestWriteOutputs:
    def test_write_json(self, tmp_path: Path) -> None:
        records = [{"title": "Test", "slug": "test"}]
        out = tmp_path / "out.json"
        write_json(records, out)
        loaded = json.loads(out.read_text())
        assert loaded == records

    def test_write_csv(self, tmp_path: Path) -> None:
        records = [
            {
                "title": "Test",
                "slug": "test",
                "ssoc_code": "12345",
                "category": "managers",
                "category_label": "Managers",
                "major_group": 1,
                "pay_monthly": 5000,
                "pay_annual": 60000,
                "pay_p25": 4000,
                "pay_p75": 6000,
                "url": "",
            }
        ]
        out = tmp_path / "out.csv"
        write_csv(records, out)
        text = out.read_text()
        assert "title,slug,ssoc_code" in text
        assert "Test,test,12345" in text
