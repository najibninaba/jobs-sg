"""Parse MOM Occupational Wage Survey 2024 into structured occupation data.

Reads wages_table4_2024.xlsx (sheet T4 — All Industries), extracts 562
occupations with gross monthly wage percentiles, assigns SSOC major groups,
and emits sg_occupations.json + sg_occupations.csv.

Usage:
    uv run python scripts/parse_wages.py
"""

import csv
import json
import re
from pathlib import Path

import openpyxl

# ── Constants ────────────────────────────────────────────────────────────────

WORKBOOK_PATH = Path("data-sources/wages_table4_2024.xlsx")
SHEET_NAME = "T4"
JSON_OUTPUT = Path("sg_occupations.json")
CSV_OUTPUT = Path("sg_occupations.csv")

EXPECTED_OCCUPATION_COUNT = 562
EXPECTED_GROUP_COUNT = 9
HEADER_SCAN_ROWS = 8  # Only scan rows 1–8 for column headers

CSV_FIELDS = [
    "title",
    "slug",
    "ssoc_code",
    "category",
    "category_label",
    "major_group",
    "pay_monthly",
    "pay_annual",
    "pay_p25",
    "pay_p75",
    "url",
]

# Canonical mapping: SSOC major group number → (category_label, category_slug)
MAJOR_GROUP_MAP: dict[int, tuple[str, str]] = {
    1: ("Managers", "managers"),
    2: ("Professionals", "professionals"),
    3: ("Associate Professionals and Technicians", "associate-professionals-technicians"),
    4: ("Clerical Support Workers", "clerical-support"),
    5: ("Service and Sales Workers", "service-sales"),
    6: ("Agricultural and Fishery Workers", "agricultural-fishery"),
    7: ("Craftsmen and Related Trades Workers", "craftsmen"),
    8: ("Plant and Machine Operators and Assemblers", "plant-machine-operators"),
    9: ("Cleaners, Labourers and Related Workers", "cleaners-labourers"),
}

# Alias table: normalized header text → major group number
_GROUP_ALIASES: dict[str, int] = {
    "MANAGERS": 1,
    "PROFESSIONALS": 2,
    "ASSOCIATE PROFESSIONALS AND TECHNICIANS": 3,
    "CLERICAL SUPPORT WORKERS": 4,
    "SERVICE AND SALES WORKERS": 5,
    "AGRICULTURAL AND FISHERY WORKERS": 6,
    "CRAFTSMEN AND RELATED TRADES WORKERS": 7,
    "PLANT AND MACHINE OPERATORS AND ASSEMBLERS": 8,
    "CLEANERS, LABOURERS AND RELATED WORKERS": 9,
    "CLEANERS LABOURERS AND RELATED WORKERS": 9,
}


# ── Helpers ──────────────────────────────────────────────────────────────────


def normalize_header(text: str) -> str:
    """Normalize major-group header text for alias matching.

    Strips leading digits/separators, collapses whitespace, uppercases,
    and normalises '&' to 'AND'.
    """
    s = str(text).strip()
    # Remove leading digit(s) and any following whitespace/separators
    s = re.sub(r"^\d+[\s.)*-]*", "", s)
    s = s.replace("&", "AND")
    s = re.sub(r"\s+", " ", s).strip().upper()
    return s


def detect_major_group(text: str) -> int | None:
    """Return the major group number if *text* is a recognized header, else None."""
    normalized = normalize_header(text)
    return _GROUP_ALIASES.get(normalized)


def parse_wage(value: object) -> int | None:
    """Parse an Excel wage cell to an integer, or None if blank/placeholder."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip().replace(",", "").replace(" ", "")
    if not s or s == "-":
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def make_slug(title: str) -> str:
    """Generate a URL-safe slug from an occupation title.

    Lowercase, '&' → 'and', strip apostrophes, non-alnum runs → '-',
    collapse hyphens, trim edges.
    """
    s = title.lower()
    s = s.replace("&", "and")
    s = s.replace("'", "")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def dedupe_slugs(records: list[dict]) -> None:
    """Mutate *records* in place to ensure unique slugs via -2, -3 suffixes."""
    seen: dict[str, int] = {}
    for rec in records:
        base = rec["slug"]
        if base not in seen:
            seen[base] = 1
        else:
            seen[base] += 1
            rec["slug"] = f"{base}-{seen[base]}"


def clean_title(text: str) -> str:
    """Collapse whitespace in an occupation title, preserving original wording."""
    return re.sub(r"\s+", " ", str(text)).strip()


def format_ssoc_code(value: object) -> str:
    """Normalise an SSOC code cell to a 5-digit zero-padded string."""
    if isinstance(value, (int, float)):
        return str(int(value)).zfill(5)
    return str(value).strip().zfill(5)


# ── Core parser ──────────────────────────────────────────────────────────────


def parse_workbook(
    workbook_path: Path = WORKBOOK_PATH,
    sheet_name: str = SHEET_NAME,
) -> list[dict]:
    """Parse the MOM wage survey workbook into a list of occupation records.

    Raises:
        FileNotFoundError: If *workbook_path* does not exist.
        ValueError: If the sheet structure is invalid or unexpected.
    """
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
        return _parse_sheet(ws)
    finally:
        wb.close()


def _parse_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """Internal: extract occupation records from a worksheet."""
    # We know the layout from exploration:
    # Col B (idx 1) = SSOC code, Col C (idx 2) = title,
    # Col D-F (idx 3-5) = basic p25/med/p75,
    # Col G-I (idx 6-8) = gross p25/med/p75
    #
    # Row 9+ = data (major group headers + occupation rows)
    COL_SSOC = 1  # 0-indexed into row tuple
    COL_TITLE = 2
    COL_GROSS_P25 = 6
    COL_GROSS_MED = 7
    COL_GROSS_P75 = 8

    records: list[dict] = []
    current_group: int | None = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
        cells = list(row)
        # Pad if row is short
        while len(cells) < 9:
            cells.append(None)

        title_cell = cells[COL_TITLE]
        gross_med = parse_wage(cells[COL_GROSS_MED])

        # Skip blank rows
        if title_cell is None and gross_med is None:
            continue

        # Check for major-group header row (ALL-CAPS title, no wage data)
        if title_cell is not None and gross_med is None:
            group = detect_major_group(str(title_cell))
            if group is not None:
                current_group = group
            continue

        # Occupation row: must have title, SSOC code, and gross median
        if title_cell is None or gross_med is None:
            continue

        ssoc_raw = cells[COL_SSOC]
        if ssoc_raw is None:
            continue

        if current_group is None:
            raise ValueError(f"Row {row_idx}: occupation found before any major-group header")

        ssoc_code = format_ssoc_code(ssoc_raw)
        title = clean_title(title_cell)
        slug = make_slug(title)
        gross_p25 = parse_wage(cells[COL_GROSS_P25])
        gross_p75 = parse_wage(cells[COL_GROSS_P75])

        category_label, category = MAJOR_GROUP_MAP[current_group]

        records.append(
            {
                "title": title,
                "slug": slug,
                "ssoc_code": ssoc_code,
                "category": category,
                "category_label": category_label,
                "major_group": current_group,
                "pay_monthly": gross_med,
                "pay_annual": gross_med * 12,
                "pay_p25": gross_p25,
                "pay_p75": gross_p75,
                "url": "",
            }
        )

    dedupe_slugs(records)
    return records


# ── Validation ───────────────────────────────────────────────────────────────


def validate_records(records: list[dict]) -> None:
    """Enforce dataset contract on parsed records.

    Raises ValueError if any check fails.
    """
    errors: list[str] = []

    # 1. Row count
    if len(records) != EXPECTED_OCCUPATION_COUNT:
        errors.append(f"Expected {EXPECTED_OCCUPATION_COUNT} occupations, got {len(records)}")

    # 2. Group count
    groups = {r["major_group"] for r in records}
    if len(groups) != EXPECTED_GROUP_COUNT:
        errors.append(
            f"Expected {EXPECTED_GROUP_COUNT} major groups, got {len(groups)}: {sorted(groups)}"
        )

    # 3. Required fields
    required_str = ["title", "slug", "ssoc_code", "category", "category_label"]
    required_int = ["pay_monthly", "pay_annual"]
    for i, r in enumerate(records):
        for field in required_str:
            if not r.get(field):
                errors.append(f"Record {i} ({r.get('title', '?')}): missing {field}")
        for field in required_int:
            if not isinstance(r.get(field), int):
                errors.append(f"Record {i} ({r.get('title', '?')}): {field} not int")

    # 4. Wage ordering (p25 <= median <= p75)
    for i, r in enumerate(records):
        p25 = r.get("pay_p25")
        med = r.get("pay_monthly")
        p75 = r.get("pay_p75")
        if isinstance(p25, int) and isinstance(med, int) and p25 > med:
            errors.append(f"Record {i} ({r['title']}): pay_p25 ({p25}) > median ({med})")
        if isinstance(med, int) and isinstance(p75, int) and med > p75:
            errors.append(f"Record {i} ({r['title']}): median ({med}) > pay_p75 ({p75})")

    # 5. Unique slugs
    slugs = [r["slug"] for r in records]
    dup_slugs = {s for s in slugs if slugs.count(s) > 1}
    if dup_slugs:
        errors.append(f"Duplicate slugs: {sorted(dup_slugs)}")

    # 6. Unique SSOC codes
    codes = [r["ssoc_code"] for r in records]
    dup_codes = {c for c in codes if codes.count(c) > 1}
    if dup_codes:
        errors.append(f"Duplicate SSOC codes: {sorted(dup_codes)}")

    # 7. SSOC code format (5 digits)
    for i, r in enumerate(records):
        if not re.match(r"^\d{5}$", r["ssoc_code"]):
            errors.append(
                f"Record {i} ({r['title']}): SSOC code '{r['ssoc_code']}' is not 5 digits"
            )

    # 8. Annual = monthly * 12
    for i, r in enumerate(records):
        if (
            isinstance(r.get("pay_monthly"), int)
            and isinstance(r.get("pay_annual"), int)
            and r["pay_annual"] != r["pay_monthly"] * 12
        ):
            errors.append(f"Record {i} ({r['title']}): pay_annual != pay_monthly * 12")

    if errors:
        msg = f"{len(errors)} validation error(s):\n" + "\n".join(f"  - {e}" for e in errors)
        raise ValueError(msg)


# ── Output ───────────────────────────────────────────────────────────────────


def write_json(records: list[dict], output_path: Path = JSON_OUTPUT) -> None:
    """Write records to a JSON file."""
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)


def write_csv(records: list[dict], output_path: Path = CSV_OUTPUT) -> None:
    """Write records to a CSV file with fixed field order."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(records)


# ── Main ─────────────────────────────────────────────────────────────────────


def main() -> None:
    """Parse, validate, and write occupation data."""
    print(f"Parsing {WORKBOOK_PATH} sheet '{SHEET_NAME}'...")
    records = parse_workbook()

    print(f"Parsed {len(records)} occupation rows")
    validate_records(records)
    print("Validation passed \u2713")

    write_json(records)
    write_csv(records)
    print(f"Wrote {JSON_OUTPUT} and {CSV_OUTPUT}")

    # Summary
    groups = sorted({r["major_group"] for r in records})
    print(f"Major groups: {len(groups)} ({groups})")
    pays = [r["pay_monthly"] for r in records]
    print(f"Gross median range: S${min(pays):,} \u2013 S${max(pays):,}")


if __name__ == "__main__":
    main()
