"""Build occupation description pages from SSOC definitions and wage data.

For each of the 562 occupations in sg_occupations.json, generates a Markdown
file in pages/<slug>.md combining SSOC 2020 definitions (primary), SSOC 2024
definitions (supplemental fill), and Singapore wage context.

These pages are consumed by score.py as LLM prompt context.

Usage:
    uv run python scripts/build_descriptions.py
"""

import argparse
import json
import re
from pathlib import Path

import openpyxl

from scripts.parse_wages import format_ssoc_code

# ── Constants ────────────────────────────────────────────────────────────

OCCUPATIONS_PATH = Path("sg_occupations.json")
SSOC2020_PATH = Path("data-sources/ssoc2020_detailed_definitions.xlsx")
SSOC2024_PATH = Path("data-sources/ssoc2024_detailed_definitions.xlsx")
OUTPUT_DIR = Path("pages")

# Column indices in both SSOC workbooks (0-indexed, header on row 5)
COL_CODE = 0
COL_TITLE = 1
COL_DEFINITION = 3
COL_TASKS = 4
COL_EXAMPLES = 6

HEADER_SCAN_ROWS = 20

# Header aliases for defensive column detection
_HEADER_ALIASES: dict[str, list[str]] = {
    "ssoc_code": ["ssoc 2020", "ssoc2020", "ssoc 2024", "ssoc2024", "ssoc code", "code"],
    "title": ["title", "occupation", "occupation title", "unit group title"],
    "definition": [
        "detailed definitions",
        "definition",
        "occupation definition",
        "description",
    ],
    "tasks": ["tasks", "main tasks", "main tasks and duties", "duties"],
    "examples": [
        "examples of job classified under this code",
        "examples",
        "examples of jobs classified here",
        "included occupations",
    ],
}


# ── Helpers ───────────────────────────────────────────────────────────────


def _normalize_header_label(text: str) -> str:
    """Normalize an Excel header cell for alias matching."""
    s = str(text).strip().lower()
    s = s.replace("&", "and")
    s = re.sub(r"\s+", " ", s)
    return s


def _clean_cell(value: object) -> str | None:
    """Clean a narrative text cell. Returns None if effectively blank."""
    if value is None:
        return None
    s = str(value).strip()
    # SSOC 2020 uses '<Blank>' as placeholder
    if not s or s == "<Blank>":
        return None
    # Clean up unusual bullet characters (SSOC 2024 uses \x9f\xa0)
    s = re.sub(r"[\x9f\xa0]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _text_to_bullets(text: str) -> str:
    """Convert a narrative text block into Markdown bullet list if splittable.

    Splits on newlines, semicolons, or bullet markers (•, -, *).
    Falls back to paragraph if fewer than 2 items result.
    """
    # Try splitting on existing bullet/line markers
    parts = re.split(r"[\n;]|(?:^|\s)[•\-\*]\s", text)
    items = [p.strip().rstrip(".") for p in parts if p.strip()]
    # Remove leading dash/bullet from items
    items = [re.sub(r"^[•\-\*]\s*", "", item).strip() for item in items]
    items = [item for item in items if item]

    if len(items) >= 2:
        return "\n".join(f"- {item}" for item in items)
    return text


# ── SSOC workbook parsing ────────────────────────────────────────────────


def find_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[int, dict[str, int]]:
    """Scan the first rows of a worksheet to find the header row.

    Returns (row_number_1indexed, column_map) where column_map maps
    canonical field names to 0-indexed column positions.

    Raises ValueError if required columns (ssoc_code, definition) are not found.
    """
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True), start=1
    ):
        cells = list(row)
        col_map: dict[str, int] = {}

        for col_idx, cell in enumerate(cells):
            if cell is None:
                continue
            normalized = _normalize_header_label(cell)
            for field, aliases in _HEADER_ALIASES.items():
                if normalized in aliases and field not in col_map:
                    col_map[field] = col_idx
                    break

        # Need at least ssoc_code and definition
        if "ssoc_code" in col_map and "definition" in col_map:
            return row_idx, col_map

    raise ValueError(
        "Could not find header row with required columns (ssoc_code, definition) "
        f"in the first {HEADER_SCAN_ROWS} rows"
    )


def load_definition_workbook(
    workbook_path: Path,
    *,
    source_year: int,
) -> dict[str, dict]:
    """Parse an SSOC definitions workbook into a code→record dict.

    Only 5-digit codes (unit-level occupations) are included.
    Returns dict keyed by normalized 5-digit SSOC code strings.
    """
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    try:
        # Try each sheet until we find one with valid headers
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                header_row, col_map = find_header_row(ws)
                return _parse_definition_rows(ws, header_row, col_map, source_year=source_year)
            except ValueError:
                continue
        raise ValueError(f"No worksheet in {workbook_path} contains required columns")
    finally:
        wb.close()


def _get_cell(field: str, row_cells: list, col_mapping: dict[str, int]) -> str | None:
    """Extract and clean a named field from a row using the column map."""
    idx = col_mapping.get(field)
    if idx is None or idx >= len(row_cells):
        return None
    return _clean_cell(row_cells[idx])


def _parse_definition_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    col_map: dict[str, int],
    *,
    source_year: int,
) -> dict[str, dict]:
    """Extract 5-digit occupation definitions from worksheet rows."""
    records: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        cells = list(row)

        # Extract and normalize code
        raw_code = cells[col_map["ssoc_code"]] if col_map["ssoc_code"] < len(cells) else None
        if raw_code is None:
            continue

        code = format_ssoc_code(raw_code)
        # Only keep 5-digit unit-level codes
        if not re.match(r"^\d{5}$", code) or len(str(raw_code).strip()) < 5:
            continue

        definition = _get_cell("definition", cells, col_map)
        tasks = _get_cell("tasks", cells, col_map)
        examples = _get_cell("examples", cells, col_map)
        title = _get_cell("title", cells, col_map)

        # Skip if no narrative content at all
        if not any([definition, tasks, examples]):
            continue

        # Prefer more complete record for duplicates
        if code in records:
            existing_fields = sum(
                1 for f in ["definition", "tasks", "examples"] if records[code].get(f)
            )
            new_fields = sum(1 for f in [definition, tasks, examples] if f)
            if new_fields <= existing_fields:
                continue

        records[code] = {
            "ssoc_code": code,
            "title": title,
            "definition": definition,
            "tasks": tasks,
            "examples": examples,
            "source_year": source_year,
        }

    return records


# ── Merge ────────────────────────────────────────────────────────────────


def merge_definitions(
    primary: dict | None,
    supplemental: dict | None,
) -> dict | None:
    """Merge SSOC 2020 (primary) and 2024 (supplemental) definition records.

    2024 only fills fields that are missing/None in 2020. Never overrides.
    """
    if primary is None and supplemental is None:
        return None
    if primary is None:
        return supplemental
    if supplemental is None:
        return primary

    merged = dict(primary)
    for field in ["definition", "tasks", "examples", "title"]:
        if not merged.get(field) and supplemental.get(field):
            merged[field] = supplemental[field]
    return merged


# ── Markdown rendering ───────────────────────────────────────────────────


def render_description(occupation: dict, definition: dict | None) -> str:
    """Render a single occupation's description as Markdown."""
    title = occupation["title"]
    ssoc_code = occupation["ssoc_code"]
    pay_monthly = occupation["pay_monthly"]
    pay_p25 = occupation.get("pay_p25")
    pay_p75 = occupation.get("pay_p75")
    category_label = occupation["category_label"]

    lines = [f"# {title} (SSOC {ssoc_code})", ""]

    # Definition
    lines.append("## Definition")
    if definition and definition.get("definition"):
        lines.append(definition["definition"])
    else:
        lines.append(
            "No official SSOC definition was found for this code. "
            "Use the occupation title and Singapore wage context below "
            "as the primary scoring context."
        )
    lines.append("")

    # Tasks
    lines.append("## Key Tasks")
    if definition and definition.get("tasks"):
        lines.append(_text_to_bullets(definition["tasks"]))
    else:
        lines.append("Not available in the SSOC definitions used for this build.")
    lines.append("")

    # Examples
    lines.append("## Examples of Jobs Classified Here")
    if definition and definition.get("examples"):
        lines.append(_text_to_bullets(definition["examples"]))
    else:
        lines.append("Not available in the SSOC definitions used for this build.")
    lines.append("")

    # Singapore Context
    lines.append("## Singapore Context")
    lines.append(f"- Monthly median gross wage: S${pay_monthly:,}")

    if isinstance(pay_p25, int) and isinstance(pay_p75, int):
        lines.append(
            f"- Pay range (25th\u201375th percentile): S${pay_p25:,} \u2013 S${pay_p75:,}"
        )
    elif isinstance(pay_p25, int):
        lines.append(f"- 25th percentile gross wage: S${pay_p25:,}")
    elif isinstance(pay_p75, int):
        lines.append(f"- 75th percentile gross wage: S${pay_p75:,}")
    else:
        lines.append("- Pay range data not available in the source wage table.")

    lines.append(f"- Major group: {category_label}")
    lines.append("")

    return "\n".join(lines)


# ── Orchestration ────────────────────────────────────────────────────────


def load_occupations(path: Path = OCCUPATIONS_PATH) -> list[dict]:
    """Load and validate occupations from JSON."""
    if not path.exists():
        raise FileNotFoundError(f"Occupations file not found: {path}")

    with open(path, encoding="utf-8") as f:
        records = json.load(f)

    if not isinstance(records, list):
        raise ValueError("Occupations JSON must be a top-level array")

    slugs_seen: set[str] = set()
    for i, r in enumerate(records):
        for field in ["title", "slug", "ssoc_code"]:
            if not r.get(field):
                raise ValueError(f"Record {i}: missing required field '{field}'")
        r["ssoc_code"] = format_ssoc_code(r["ssoc_code"])
        if r["slug"] in slugs_seen:
            raise ValueError(f"Record {i}: duplicate slug '{r['slug']}'")
        slugs_seen.add(r["slug"])

    return records


def write_descriptions(
    occupations: list[dict],
    defs_2020: dict[str, dict],
    defs_2024: dict[str, dict],
    output_dir: Path = OUTPUT_DIR,
) -> dict[str, int]:
    """Generate Markdown description files for all occupations.

    Returns coverage stats.
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    stats = {"written": 0, "with_2020": 0, "with_2024_fill": 0, "fallback_only": 0}

    for occ in occupations:
        code = occ["ssoc_code"]
        primary = defs_2020.get(code)
        supplemental = defs_2024.get(code)
        merged = merge_definitions(primary, supplemental)

        # Track coverage stats
        if primary:
            stats["with_2020"] += 1
            if supplemental and any(
                not primary.get(f) and supplemental.get(f)
                for f in ["definition", "tasks", "examples"]
            ):
                stats["with_2024_fill"] += 1
        elif supplemental:
            stats["with_2024_fill"] += 1
        else:
            stats["fallback_only"] += 1

        md = render_description(occ, merged)
        out_path = output_dir / f"{occ['slug']}.md"
        out_path.write_text(md, encoding="utf-8")
        stats["written"] += 1

    return stats


def main(argv: list[str] | None = None) -> None:
    """Parse SSOC definitions, merge with occupations, write pages."""
    parser = argparse.ArgumentParser(description="Build occupation description pages")
    parser.add_argument(
        "--occupations", type=Path, default=OCCUPATIONS_PATH, help="Path to sg_occupations.json"
    )
    parser.add_argument(
        "--ssoc2020", type=Path, default=SSOC2020_PATH, help="SSOC 2020 definitions workbook"
    )
    parser.add_argument(
        "--ssoc2024", type=Path, default=SSOC2024_PATH, help="SSOC 2024 definitions workbook"
    )
    parser.add_argument(
        "--output-dir", type=Path, default=OUTPUT_DIR, help="Output directory for .md files"
    )
    args = parser.parse_args(argv)

    occupations = load_occupations(args.occupations)
    print(f"Loaded {len(occupations)} occupations")

    print(f"Parsing SSOC 2020 definitions from {args.ssoc2020}...")
    defs_2020 = load_definition_workbook(args.ssoc2020, source_year=2020)
    print(f"  Found {len(defs_2020)} 5-digit definitions")

    print(f"Parsing SSOC 2024 definitions from {args.ssoc2024}...")
    defs_2024 = load_definition_workbook(args.ssoc2024, source_year=2024)
    print(f"  Found {len(defs_2024)} 5-digit definitions")

    stats = write_descriptions(occupations, defs_2020, defs_2024, args.output_dir)
    print(f"\nWrote {stats['written']} pages to {args.output_dir}/")
    print(f"  With SSOC 2020 definition: {stats['with_2020']}")
    print(f"  With SSOC 2024 fill: {stats['with_2024_fill']}")
    print(f"  Fallback only (no definition): {stats['fallback_only']}")


if __name__ == "__main__":
    main()
